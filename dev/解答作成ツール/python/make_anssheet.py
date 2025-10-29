# -*- coding: utf-8 -*-
import openpyxl

def find_first_blank_row(sheet, column):
    """
    指定されたシートの特定の列を2行目から順に見ていき、
    最初に空白が見つかった行の番号を返します。
    """
    row = 2
    while sheet[f"{column}{row}"].value is not None and sheet[f"{column}{row}"].value != '':
        row += 1
    return row

def load_constraint_pairs(sheet, col1='A', col2='B'):
    """
    シートから制約となる整数のペアを読み込み、セットとして返します。
    A列が空になるまで読み込みを続けます。
    """
    pairs = set()
    row = 1
    while True:
        val1 = sheet[f"{col1}{row}"].value
        val2 = sheet[f"{col2}{row}"].value

        if val1 is None or str(val1).strip() == '':
            break
        
        try:
            pairs.add((int(val1), int(val2)))
        except (ValueError, TypeError):
            print(f"警告: シート '{sheet.title}' の {row} 行目の値 ({val1}, {val2}) は整数ペアとして解釈できませんでした。")
        
        row += 1
    return pairs

def load_existing_ans_data(sheet):
    """
    Ansシートから既存のデータ（A, B, C列の組み合わせ）を読み込み、
    セットとして返します。
    """
    existing_data = set()
    row = 2  # ヘッダー行をスキップ
    while True:
        val_a = sheet[f"A{row}"].value
        val_b = sheet[f"B{row}"].value
        val_c = sheet[f"C{row}"].value
        
        # いずれかの値がNoneまたは空の場合は終了
        if val_a is None or val_b is None or val_c is None:
            break
        if str(val_a).strip() == '' or str(val_b).strip() == '' or str(val_c).strip() == '':
            break
        
        try:
            existing_data.add((int(val_a), int(val_b), int(val_c)))
        except (ValueError, TypeError):
            print(f"警告: Ansシートの {row} 行目の値 ({val_a}, {val_b}, {val_c}) は整数として解釈できませんでした。")
        
        row += 1
    
    return existing_data

def generate_combinations():
    """
    Excelファイルからデータを読み込み、制約条件でフィルタリングした
    全組み合わせをテキストファイルと'Ans'シートに出力する関数。
    """
    # --- 設定項目 ---
    input_excel_file = 'スロット.xlsx'
    output_txt_file = 'Problems.txt'
    ans_sheet_name = 'Ans'
    target_sheets = {
        'L': 'Que_L',
        'C': 'Que_C',
        'R': 'Que_R'
    }
    constraint_sheets = {
        'LC': 'LC',
        'LR': 'LR',
        'CR': 'CR'
    }
    target_column = 'B'
    id_column = 'A'

    try:
        workbook = openpyxl.load_workbook(input_excel_file)
    except FileNotFoundError:
        print(f"エラー: ファイル '{input_excel_file}' が見つかりません。スクリプトと同じフォルダに配置してください。")
        return
    except Exception as e:
        print(f"エラー: Excelファイルを開けませんでした: {e}")
        return

    try:
        if ans_sheet_name not in workbook.sheetnames:
            print(f"エラー: シート '{ans_sheet_name}' が見つかりません。処理を中断します。")
            return
        ans_sheet = workbook[ans_sheet_name]
        
        existing_ans_data = load_existing_ans_data(ans_sheet)
        print(f"--- Ansシートの既存データを読み込みました ---")
        print(f"- 既存の組み合わせ数: {len(existing_ans_data)} 個")

    except Exception as e:
        print(f"エラー: '{ans_sheet_name}' シートの準備中にエラーが発生しました: {e}")
        return

    try:
        sheet_LC = workbook[constraint_sheets['LC']]
        sheet_LR = workbook[constraint_sheets['LR']]
        sheet_CR = workbook[constraint_sheets['CR']]

        lc_pairs = load_constraint_pairs(sheet_LC)
        lr_pairs = load_constraint_pairs(sheet_LR)
        cr_pairs = load_constraint_pairs(sheet_CR)
        
        print("--- 制約情報を読み込みました ---")
        print(f"- {constraint_sheets['LC']}シートから {len(lc_pairs)} 個のペア")
        print(f"- {constraint_sheets['LR']}シートから {len(lr_pairs)} 個のペア")
        print(f"- {constraint_sheets['CR']}シートから {len(cr_pairs)} 個のペア")

    except KeyError as e:
        print(f"エラー: 制約シート '{e.args[0]}' がファイル内に見つかりません。")
        return

    try:
        sheet_L = workbook[target_sheets['L']]
        sheet_C = workbook[target_sheets['C']]
        sheet_R = workbook[target_sheets['R']]

        Ls = find_first_blank_row(sheet_L, target_column)
        Cs = find_first_blank_row(sheet_C, target_column)
        Rs = find_first_blank_row(sheet_R, target_column)
    except KeyError as e:
        print(f"エラー: シート '{e.args[0]}' がファイル内に見つかりません。")
        return


    output_count = 0
    ans_row = 2 + len(existing_ans_data)
    print(f"Ansシートの書き込み開始行: {ans_row}")
    
    with open(output_txt_file, 'a', encoding='utf-8') as f:
        for i in range(2, Ls):
            for j in range(2, Cs):
                for k in range(2, Rs):
                    
                    try:
                        type_R = sheet_L[f'{id_column}{i}'].value
                        type_C = sheet_C[f'{id_column}{j}'].value
                        type_L = sheet_R[f'{id_column}{k}'].value

                        row_R = i
                        row_C = j
                        row_L = k

                        if type_R is None or type_C is None or type_L is None:
                            continue 
                        
                        type_R, type_C, type_L = int(type_R), int(type_C), int(type_L)
                    except (ValueError, TypeError):
                        continue

                    if (type_R, type_C) in lc_pairs and \
                       (type_R, type_L) in lr_pairs and \
                       (type_C, type_L) in cr_pairs:
                        
                        if (row_R, row_C, row_L) not in existing_ans_data:
                            val_L = sheet_L[f'{target_column}{i}'].value or ''
                            val_C = sheet_C[f'{target_column}{j}'].value or ''
                            val_R = sheet_R[f'{target_column}{k}'].value or ''
                            concatenated_string = str(val_L) + str(val_C) + str(val_R)
                            
                            output_line = f"{row_R} {row_C} {row_L} {concatenated_string}\n"
                            f.write(output_line)

                            ans_sheet.cell(row=ans_row, column=1, value=row_R)
                            ans_sheet.cell(row=ans_row, column=2, value=row_C)
                            ans_sheet.cell(row=ans_row, column=3, value=row_L)
                            ans_sheet.cell(row=ans_row, column=9, value=concatenated_string)
                            
                            output_count += 1
                            ans_row += 1

    try:
        workbook.save(input_excel_file)
        print(f"\n'{input_excel_file}' の '{ans_sheet_name}' シートを更新しました。")
    except Exception as e:
        print(f"\nエラー: Excelファイル '{input_excel_file}' の保存中にエラーが発生しました。ファイルが開かれている可能性があります。")
        print(f"詳細: {e}")

    print("\n--- 処理完了 ---")
    print(f"'{output_txt_file}' に {output_count} 件の新規組み合わせを出力しました。")
    print(f"（既存データとの重複を除外）")
    print("\n処理範囲:")
    print(f"- {target_sheets['L']}: {target_column}列の {Ls - 1} 行目まで")
    print(f"- {target_sheets['C']}: {target_column}列の {Cs - 1} 行目まで")
    print(f"- {target_sheets['R']}: {target_column}列の {Rs - 1} 行目まで")


if __name__ == "__main__":
    generate_combinations()
