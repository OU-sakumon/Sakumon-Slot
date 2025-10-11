# -*- coding: utf-8 -*-
import openpyxl
import json
import sys

def find_first_blank_row(sheet):
    """
    指定されたシートのD,E,F,G列を2行目から順に見ていき、
    最初にすべての列が空白の行の番号を返します。
    """
    row = 2
    while True:
        d_value = sheet[f"D{row}"].value
        e_value = sheet[f"E{row}"].value
        f_value = sheet[f"F{row}"].value
        g_value = sheet[f"G{row}"].value
        
        # D,E,F,G列すべてが空白またはNoneの場合
        if (d_value is None or d_value == '') and \
           (e_value is None or e_value == '') and \
           (f_value is None or f_value == '') and \
           (g_value is None or g_value == ''):
            return row
        
        row += 1

def parse_json_to_columns(json_data):
    """
    JSONデータをパースしてD,E,F,G列用のデータに変換します。
    期待されるJSON形式: {"D": value1, "E": value2, "F": value3, "G": value4}
    """
    try:
        if isinstance(json_data, str):
            parsed = json.loads(json_data)
        else:
            parsed = json_data
        
        d_value = parsed.get('D', '')
        e_value = parsed.get('E', '')
        f_value = parsed.get('F', '')
        g_value = parsed.get('G', '')
        
        return d_value, e_value, f_value, g_value
        
    except json.JSONDecodeError as e:
        print(f"JSON解析エラー: {e}")
        return None, None, None, None
    except Exception as e:
        print(f"データ処理エラー: {e}")
        return None, None, None, None

def import_json_to_ans_sheet(json_input, excel_file='スロット.xlsx', ans_sheet_name='Ans'):
    """
    JSONデータをAnsシートのD,E,F,G列にインポートします。
    
    Args:
        json_input: JSON文字列、辞書、またはJSONファイルのパス
        excel_file: 対象のExcelファイル名
        ans_sheet_name: 対象のシート名
    """
    
    if isinstance(json_input, str):
        if json_input.endswith('.json'):
            try:
                with open(json_input, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)
            except FileNotFoundError:
                print(f"エラー: ファイル '{json_input}' が見つかりません。")
                return
            except Exception as e:
                print(f"エラー: ファイル読み込み中にエラーが発生しました: {e}")
                return
        else:
            json_data = json_input
    else:
        json_data = json_input
    
    try:
        workbook = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        print(f"エラー: ファイル '{excel_file}' が見つかりません。")
        return
    except Exception as e:
        print(f"エラー: Excelファイルを開けませんでした: {e}")
        return
    
    try:
        if ans_sheet_name not in workbook.sheetnames:
            print(f"エラー: シート '{ans_sheet_name}' が見つかりません。")
            return
        ans_sheet = workbook[ans_sheet_name]
    except Exception as e:
        print(f"エラー: '{ans_sheet_name}' シートの取得中にエラーが発生しました: {e}")
        return
    
    start_row = find_first_blank_row(ans_sheet)
    
    if isinstance(json_data, list):
        success_count = 0
        error_count = 0
        
        for i, item in enumerate(json_data):
            d_val, e_val, f_val, g_val = parse_json_to_columns(item)
            
            if d_val is not None:
                ans_sheet.cell(row=start_row, column=4, value=d_val)
                ans_sheet.cell(row=start_row, column=5, value=e_val)
                ans_sheet.cell(row=start_row, column=6, value=f_val)
                ans_sheet.cell(row=start_row, column=7, value=g_val)
                
                success_count += 1
                start_row += 1
            else:
                error_count += 1
                print(f"警告: {i+1}番目のJSONアイテムの解析に失敗しました。スキップします。")
        
        print(f"処理完了: 成功 {success_count} 件, エラー {error_count} 件")
        
    else:
        d_val, e_val, f_val, g_val = parse_json_to_columns(json_data)
        
        if d_val is not None:
            ans_sheet.cell(row=start_row, column=4, value=d_val)
            ans_sheet.cell(row=start_row, column=5, value=e_val)
            ans_sheet.cell(row=start_row, column=6, value=f_val)
            ans_sheet.cell(row=start_row, column=7, value=g_val)
            
            print(f"JSONデータを {start_row} 行目に書き込みました。")
        else:
            print("エラー: JSONデータの解析に失敗しました。")
            return
    
    try:
        workbook.save(excel_file)
        print(f"'{excel_file}' の '{ans_sheet_name}' シートを更新しました。")
    except Exception as e:
        print(f"エラー: Excelファイルの保存中にエラーが発生しました: {e}")

def main():
    """
    コマンドライン引数からJSONデータを受け取って処理します。
    使用方法:
    python import_answers.py
    python import_answers.py '{"D": "value1", "E": "value2", "F": "value3", "G": "value4"}'
    または
    python import_answers.py data.json
    """
    if len(sys.argv) < 2:
        # 引数がない場合はデフォルトでanswers.jsonを使用
        json_input = 'answers.json'
    else:
        json_input = sys.argv[1]
    
    import_json_to_ans_sheet(json_input)

if __name__ == "__main__":
    main()
